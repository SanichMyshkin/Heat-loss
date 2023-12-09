import sys
import os
import openpyxl

from PyQt6.QtWidgets import QMainWindow, QHeaderView, \
    QApplication, QFileDialog, QMessageBox, QTableWidgetItem
from PyQt6 import uic, QtGui

from materials import MaterialsTable
from walls import WallsTable
from rooms import RoomsTable
from calculate import CalculateTable

ui_path = os.path.abspath(os.path.join(os.path.dirname(__file__), 'gui.ui'))


class GUI(QMainWindow, MaterialsTable,
          WallsTable, RoomsTable, CalculateTable):

    def __init__(self):
        super(GUI, self).__init__()
        uic.loadUi(ui_path, self)

        self.setWindowTitle(
            "Расчет тепловых потерь через внутренние ограждающие конструкции")
        self.setWindowIcon(QtGui.QIcon('icon.png'))
        # Списки для передачи и хранения материалов
        self.list_materials = []
        self.list_material_position = []

        # Списки для для передачи и обновления помещений
        self.list_rooms = []
        self.list_rooms_position = []

        # Списки для передачи и обновлений стен
        self.list_walls = []
        self.list_walls_position = []

        # Инициализация методов для каждой таблицы
        self.set_autosize_table()
        self.materials()
        self.calculate()
        self.rooms()
        self.walls()

        # Иницализация методов для очистки данных
        self.ClearMaterials.triggered.connect(self.ClearTableMaterials)
        self.ClearWalls.triggered.connect(self.ClearTableWalls)
        self.ClearRooms.triggered.connect(self.ClearTableRooms)
        self.ClearCalculate.triggered.connect(self.ClearTableCalculate)
        self.ClearAll.triggered.connect(self.ClearAllTable)

        # Инициализация методов для Сохранения и экспорта данных
        self.SaveAll.triggered.connect(self.Save)
        self.ExpMaterialsAction.triggered.connect(self.ExpMaterials)
        self.ExpRoomsAction.triggered.connect(self.ExpRooms)

        # Инициализация методов для Имопрта Данных
        self.ImportMaterialsPushButton.clicked.connect(
            self.LoadMaterials)
        self.ImportRoomsPushButton.clicked.connect(
            self.LoadRooms)

        # Иницализация методов справки
        self.AuthorsAction.triggered.connect(self.AuthorsMessage)
        self.TutorialAction.triggered.connect(self.TutorialMessage)

    def TutorialMessage(self):
        message_box = QMessageBox()
        message_box.information(
            None, "Руководство", "Тут будет инструкция по использованию приложением\n (◕‿◕)")
        message_box.setFixedSize(500, 200)

    def AuthorsMessage(self):
        message_box = QMessageBox()
        message_box.information(
            None, "Авторы",
            "Данное приложение было разработано студентами НИУ МГСУ 4 курса направления ИСТАС, \nгруппы ИЦТМС 4-2, \n\nМышкиным А.В. \nи \nДубровиным В.А.\n2023 г.")
        message_box.setFixedSize(500, 200)

    def ClearAllTable(self):
        '''Метод очищения всех таблиц'''
        self.ClearTableMaterials()
        self.ClearTableWalls()
        self.ClearTableRooms()
        self.ClearTableCalculate()

    def set_autosize_table(self):
        '''Расстягивание ширины заголовков таблиц под ширину экрана'''
        self.MaterialsTableWidget.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)

        self.RoomsTableWidget.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)

        self.WallsTableWidget.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)

        self.CalculateTableWidget.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)

    def materials(self):
        '''Методы для таблицы материалов'''
        self.AddRowMaterialsPushButton.clicked.connect(self.AddMaterialsRow)
        self.RemoveRowMaterialsPushButton.clicked.connect(self.RemoveMaterialsRow)  # noqa E501
        self.MaterialsCountSpinBox()

    def rooms(self):
        '''Методы для таблицы помещение'''
        self.AddRowRoomsPushButton.clicked.connect(self.AddRoomsRow)
        self.RemoveRowRoomsPushButton.clicked.connect(self.RemoveRoomsRow)  # noqa E501
        self.RoomsCountSpinBox()

    def calculate(self):
        '''Методы для таблицы расчеты'''
        self.AddRowCalculatePushButton.clicked.connect(self.AddCalculateRow)
        self.RemoveRowCalculatePushButton.clicked.connect(self.RemoveCalculateRow)  # noqa E501
        self.GetRoomsPushButton.clicked.connect(self.GetRoomsAndWalls)
        self.ShowCalculatePushButton.clicked.connect(self.ShowCalculate)

        self.CalculateCountSpinBox()
        self.CalculateRoomsComboBox()
        self.CalculateWallsComboBox()
        self.MakeCalculateCellReadOnly()

    def walls(self):
        '''Методы для таблицы стен'''
        self.AddRowWallsPushButton.clicked.connect(self.AddWallsRow)
        self.RemoveRowWallsPushButton.clicked.connect(self.RemoveWallsRow)
        self.GetMaterialsPushButton.clicked.connect(self.GetMaterials)
        self.CalculateWallsPushButton.clicked.connect(self.CalculateWalls)  # noqa
        self.WallLayerComboBox()
        self.MakeWallsCellReadOnly()

    def LoadMaterials(self):
        file_dialog = QFileDialog()
        path, _ = file_dialog.getOpenFileName(
            self, "Открыть файл", "", "Excel Files (*.xlsx);;All Files (*)")
        if not path:
            return

        if not os.path.getsize(path):
            QMessageBox.warning(self, "Внимание", "Выбранный файл пуст.")
            return

        try:
            wb = openpyxl.load_workbook(path)
            if any(sheet.max_row > 1 for sheet in wb.worksheets):
                self.ImportMaterials(wb)
                QMessageBox.information(
                    self, "Успешно", "Данные успешно загружены!")
            else:
                QMessageBox.warning(
                    self, "Внимание", "Выбранный файл Excel пуст.")
        except Exception as e:
            QMessageBox.warning(
                self, "Ошибка", f"Произошла ошибка при загрузке файла: {str(e)}")

    def ImportMaterials(self, workbook):
        ws_materials = workbook.active
        self.MaterialsTableWidget.setRowCount(0)

        for row in range(2, ws_materials.max_row + 1):
            name_materials = ws_materials.cell(row=row, column=1).value
            coeff_value = ws_materials.cell(row=row, column=2).value
            thickness_value = ws_materials.cell(row=row, column=3).value

            self.AddMaterialsRow()
            self.MaterialsTableWidget.setItem(
                row - 2, 0, QTableWidgetItem(name_materials))
            self.MaterialsTableWidget.cellWidget(row - 2, 1).setValue(
                0.0 if coeff_value is None or isinstance(coeff_value, str) else coeff_value)
            self.MaterialsTableWidget.cellWidget(row - 2, 2).setValue(
                0.0 if thickness_value is None or isinstance(thickness_value, str) else thickness_value)

    def LoadRooms(self):
        file_dialog = QFileDialog()
        path, _ = file_dialog.getOpenFileName(
            self, "Открыть файл", "", "Excel Files (*.xlsx);;All Files (*)")
        if not path:
            return

        if not os.path.getsize(path):
            QMessageBox.warning(self, "Внимание", "Выбранный файл пуст.")
            return

        try:
            wb = openpyxl.load_workbook(path)
            if any(sheet.max_row > 1 for sheet in wb.worksheets):
                self.ImportRooms(wb)
                QMessageBox.information(
                    self, "Успешно", "Данные успешно загружены!")
            else:
                QMessageBox.warning(
                    self, "Внимание", "Выбранный файл Excel пуст.")
        except Exception as e:
            QMessageBox.warning(
                self, "Ошибка", f"Произошла ошибка при загрузке файла: {str(e)}")

    def ImportRooms(self, workbook):
        ws_rooms = workbook.active
        self.RoomsTableWidget.setRowCount(0)

        for row in range(2, ws_rooms.max_row + 1):
            name_room = ws_rooms.cell(row=row, column=1).value
            temp = ws_rooms.cell(row=row, column=2).value

            self.AddRoomsRow()
            self.RoomsTableWidget.setItem(
                row - 2, 0, QTableWidgetItem(name_room))
            self.RoomsTableWidget.cellWidget(row - 2, 1).setValue(
                0.0 if temp is None or isinstance(temp, str) else temp)

    def Save(self):
        file_dialog = QFileDialog()
        path, _ = file_dialog.getSaveFileName(
            self, "Сохранить файл", "Расчеты", "Excel Files (*.xlsx);;All Files (*)")
        if not path:
            return
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        self.SaveMaterials(wb)
        self.SaveRooms(wb)
        self.SaveWalls(wb)
        self.SaveCalculate(wb)
        wb.save(path)
        QMessageBox.information(
            self, "Успешно", "Файл успешно сохранен!")

    def SaveRooms(self, workbook):
        ws_rooms = workbook.create_sheet("Помещения", 0)
        column_headers = ["Название", "Температура"]
        for col_num, header in enumerate(column_headers, 1):
            header_cell = ws_rooms.cell(row=1, column=col_num)
            header_cell.value = header
        for row in range(self.RoomsTableWidget.rowCount()):
            name_rooms_item = self.RoomsTableWidget.item(row, 0)
            temperature_value = self.RoomsTableWidget.cellWidget(
                row, 1).value()
            name_rooms = name_rooms_item.text() if name_rooms_item else "-"
            ws_rooms.cell(row=row + 2, column=1, value=name_rooms)
            ws_rooms.cell(row=row + 2, column=2, value=temperature_value)

    def SaveMaterials(self, workbook):
        ws_materials = workbook.create_sheet("Материалы", 1)

        column_headers = ["Наименование", "λ", "δ,м"]
        for col_num, header in enumerate(column_headers, 1):
            header_cell = ws_materials.cell(row=1, column=col_num)
            header_cell.value = header

        for row in range(self.MaterialsTableWidget.rowCount()):
            name_materials_item = self.MaterialsTableWidget.item(row, 0)
            coeff_value = self.MaterialsTableWidget.cellWidget(row, 1).value()
            thickness_value = self.MaterialsTableWidget.cellWidget(
                row, 2).value()
            name_materials = name_materials_item.text() if name_materials_item else ""
            ws_materials.cell(row=row + 2, column=1, value=name_materials)
            ws_materials.cell(row=row + 2, column=2, value=coeff_value)
            ws_materials.cell(row=row + 2, column=3, value=thickness_value)

    def SaveWalls(self, workbook):
        ws_walls = workbook.create_sheet("Стены", 2)
        column_headers = ["Наименование", "Слой №1", "Слой №2",
                          "Слой №3", "Слой №4", "Слой №5", "δ/λ", "R"]
        for col_num, header in enumerate(column_headers, 1):
            header_cell = ws_walls.cell(row=1, column=col_num)
            header_cell.value = header

        for row in range(self.WallsTableWidget.rowCount()):
            name_walls_item = self.WallsTableWidget.item(row, 0)
            layer_1_value = self.WallsTableWidget.cellWidget(
                row, 1).currentText()
            layer_2_value = self.WallsTableWidget.cellWidget(
                row, 2).currentText()
            layer_3_value = self.WallsTableWidget.cellWidget(
                row, 3).currentText()
            layer_4_value = self.WallsTableWidget.cellWidget(
                row, 4).currentText()
            layer_5_value = self.WallsTableWidget.cellWidget(
                row, 5).currentText()

            coeff_1_value = self.WallsTableWidget.item(row, 6).text()
            coeff_2_value = self.WallsTableWidget.item(row, 7).text()

            ws_walls.cell(row=row + 2, column=1,
                          value=name_walls_item.text() if name_walls_item else "")
            ws_walls.cell(row=row + 2, column=2, value=layer_1_value)
            ws_walls.cell(row=row + 2, column=3, value=layer_2_value)
            ws_walls.cell(row=row + 2, column=4, value=layer_3_value)
            ws_walls.cell(row=row + 2, column=5, value=layer_4_value)
            ws_walls.cell(row=row + 2, column=6, value=layer_5_value)

            ws_walls.cell(row=row + 2, column=7,
                          value=coeff_1_value if coeff_1_value else "")
            ws_walls.cell(row=row + 2, column=8,
                          value=coeff_2_value if coeff_2_value else "")

    def SaveCalculate(self, workbook):
        ws_calc = workbook.create_sheet("Итоговые расчеты", 3)
        column_headers = ["Помещение №1", "Помещение №2",
                          "t1, °С", "t2, °С",
                          "Стена", "Высота", "Ширрина", "Площадь",
                          "К, Вт/(м2 ˚С)", "Теплопотери"]
        for col_num, header in enumerate(column_headers, 1):
            header_cell = ws_calc.cell(row=1, column=col_num)
            header_cell.value = header

        for row in range(self.CalculateTableWidget.rowCount()):
            name_wall_combobox1 = self.CalculateTableWidget.cellWidget(row, 0)
            name_wall_combobox2 = self.CalculateTableWidget.cellWidget(row, 1)
            temp_1_item = self.CalculateTableWidget.item(row, 2)
            temp_2_item = self.CalculateTableWidget.item(row, 3)
            wall_combobox = self.CalculateTableWidget.cellWidget(row, 4)
            width_spinbox = self.CalculateTableWidget.cellWidget(row, 6)
            height_spinbox = self.CalculateTableWidget.cellWidget(row, 5)
            area_item = self.CalculateTableWidget.item(row, 7)
            coeff_1_item = self.CalculateTableWidget.item(row, 8)
            coeff_2_item = self.CalculateTableWidget.item(row, 9)

            ws_calc.cell(row=row + 2, column=1,
                         value=name_wall_combobox1.currentText() if name_wall_combobox1 else "")

            ws_calc.cell(row=row + 2, column=2,
                         value=name_wall_combobox2.currentText() if name_wall_combobox2 else "")

            ws_calc.cell(row=row + 2, column=3,
                         value=temp_1_item.text() if temp_1_item else "")

            ws_calc.cell(row=row + 2, column=4,
                         value=temp_2_item.text() if temp_2_item else "")

            ws_calc.cell(row=row + 2, column=5,
                         value=wall_combobox.currentText() if wall_combobox else "")

            ws_calc.cell(row=row + 2, column=6,
                         value=width_spinbox.value() if width_spinbox else 0)

            ws_calc.cell(row=row + 2, column=7,
                         value=height_spinbox.value() if height_spinbox else 0)

            ws_calc.cell(row=row + 2, column=8,
                         value=area_item.text() if area_item else "")

            ws_calc.cell(row=row + 2, column=9,
                         value=coeff_1_item.text() if coeff_1_item else "")

            ws_calc.cell(row=row + 2, column=10,
                         value=coeff_2_item.text() if coeff_2_item else "")

    def ExpMaterials(self):
        file_dialog = QFileDialog()
        path, _ = file_dialog.getSaveFileName(
            self, "Экспорт материалов", "Материалы", "Excel Files (*.xlsx);;All Files (*)")
        if not path:
            return
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        self.SaveMaterials(wb)
        wb.save(path)
        QMessageBox.information(
            self, "Успешно", "Материалы успешно экспортированны!")

    def ExpRooms(self):
        file_dialog = QFileDialog()
        path, _ = file_dialog.getSaveFileName(
            self, "Экспорт помещений", "Помещения", "Excel Files (*.xlsx);;All Files (*)")
        if not path:
            return
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        self.SaveRooms(wb)
        wb.save(path)
        QMessageBox.information(
            self, "Успешно", "Помещения успешно экспортированны!")


def main():
    app = QApplication(sys.argv)
    app.setWindowIcon(QtGui.QIcon('icon.png'))
    window = GUI()
    window.show()
    app.exec()


if __name__ == '__main__':
    main()
